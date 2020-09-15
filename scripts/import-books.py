# coding=utf-8
__author__ = 'defence.zhang@gmail.com'
__date__ = "2020/09/14 上午11:38:00"

import sys
reload(sys)
sys.setdefaultencoding('utf-8')

import os
# import chardet
from openpyxl import load_workbook
import pymysql
from pymysql.cursors import DictCursor

workbook = load_workbook(os.path.join(os.path.dirname(__file__), "books.xlsx"))
sheet1 = workbook.worksheets[0]

print sheet1[1][0].value
# print sheet1.cell(1, 1).value

conn = pymysql.connect(autocommit=True, host="localhost", user="web", passwd="qaz123", db="child_library")
cursor = conn.cursor()

nameErrors = []
authorErrors = []
errors = []
for (idx, line) in enumerate(sheet1):
    if idx == 0: continue
    code = line[0].value
    name = line[1].value
    isbn = line[2].value
    price = line[3].value
    press = line[4].value
    status = line[5].value
    _lib = line[6].value
    created = line[7].value
    author = line[8].value
    intro = line[9].value
    pubDate = line[10].value
    lang = line[11].value
    remark = line[12].value
    try: print name
    except UnicodeEncodeError:
        nameErrors.append('%s: %s' % (idx, code))
        name = None
        # name = name.encode('utf-8')
        # print 'ERROR:', idx, code, 'name'
    try: print author
    except UnicodeEncodeError:
        authorErrors.append('%s: %s' % (idx, code))
        author = None
        # author = author.encode('utf-8')
        # print 'ERROR:', idx, code, 'author'
    # try: print idx, 'AfterEncoder:', name, author
    # except:
    #     errors.append('%s: %s' % (idx, code))
    #     continue

    # print name
    # print isbn
    # print price
    # print press
    # print status
    # print _lib
    # print created
    # print author
    # print intro
    # print pubDate
    # print lang
    # print remark
    try:
        print idx, code, name, isbn, price, press, status, _lib, created, author, intro, pubDate, lang, remark
    except:
        print idx, code, '-' * 20
        errors.append('%s: %s' % (idx, code))
        continue
        # print idx, code, name.decode('gbk').encode('utf-8'), author.decode('gbk').encode('utf-8')
    try:
        cursor.execute('INSERT INTO books (isbn, code, name, price, press, created, author, pubDate, intro, remark, lang) VALUE(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);',
            (isbn, code, name, price or None, press, created, author, pubDate, intro, remark, lang))
    except:
        errors.append('%s: %s' % (idx, code))
        continue

print 'nameErrors', nameErrors
print 'authorErrors', authorErrors
print 'Errors', errors

conn.close()


"""
nameErrors [u'371: 38630003818', u'388: 38630004129', u'487: 38630004938', u'599: 38630005775', u'600: 38630005768', u'601: 38630005782', u'617: 38630005843', u'845: 38630009681', u'1143: 38630011493', u'1217: 38630011592', u'1218: 38630011547', u'1280: 38630012445', u'1334: 38630014470', u'1765: 38630016917', u'1776: 
38630015835', u'1790: 38630015750', u'1848: 38630015996', u'1987: 38630019475']

authorErrors [u'149: 38630001340', u'153: 38630001418', u'198: 38630001784', u'199: 21294273', u'279: 38630002583', u'361: 38630011622', u'413: 38630004280', u'501: 38630005003', u'502: 07271298', u'563: 38630005546', u'564: 05124176', u'574: 38630005669', u'576: 38630005676', u'577: 38630005645', u'585: 38630005720', 
u'1047: 38630013046',
// 以下作者信息待修正
u'1080: 38630012315', u'1107: 38630012681', u'1132: 38630012094', u'1146: 38630011387', u'1149: 38630010403', u'1155: 38630011998', u'1160: 38630012032', u'1191: 38630010359', u'1214: 38630011554', u'1241: 38630011271', u'1246: 38630011370', u'1247: 38630011349', u'1280: 38630012445', u'1282: 38630012407', u'1334: 38630014470', u'1337: 38630013497', u'1360: 38630014296', u'1371: 38630013473', u'1379: 38630013398', u'1424: 38630014173', u'1429: 38630014135', u'1435: 38630016344', u'1441: 38630015644', u'1464: 38630018126', u'1509: 38630013800', u'1512: 38630013787', u'1513: 38630016573', u'1549: 38630018096', u'1553: 38630018065', u'1555: 38630018058', u'1603: 38630015088', u'1610: 38630015040', u'1614: 38630015279', u'1623: 38630017471', u'1682: 38630019451', u'1717: 
38630019772', u'1723: 38630019758', u'1738: 38630019666', u'1743: 38630017402', u'1751: 38630019611', u'1761: 38630019536', u'1764: 38630019505', u'1772: 38630016894', u'1778: 38630019307', u'1786: 38630015774', u'1794: 38630015729', u'1798: 38630015705', u'1803: 38630019185', u'1843: 38630016030', u'1903: 38630018393', u'1912: 38630018348', u'1942: 38630018171', u'1957: 38630019031', u'1959: 38630017778']

Errors [u'486: 38630004914', u'663: 38630009582']
"""
